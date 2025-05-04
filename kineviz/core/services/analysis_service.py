import tempfile
import shutil # Añadir para borrar directorio
import logging
from pathlib import Path
from datetime import datetime
import numpy as np
import pandas as pd

# Importar servicios y helpers
from .file_service import FileService
from .study_service import StudyService
from kineviz.ui.widgets import charting  # Importar nuestro módulo de gráficos
# from kineviz.core.data_processing import file_handlers # No usado directamente aquí
# Importar el validador de nombres de archivo
from kineviz.ui.utils.validators import validate_filename_for_study_criteria

# Ya no se necesita AppSettings
# from kineviz.config.settings import AppSettings

# Importar reportlab
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Image,
                                Table, TableStyle)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
# import itertools # No usado directamente aquí
import json
import time # Añadir import


# Define logger before it's potentially used in the except block below
logger = logging.getLogger(__name__)

# Importar scipy para tests estadísticos
try:
    from scipy import stats
except ImportError:
    logger.warning("Scipy no está instalado. Las pruebas estadísticas no estarán disponibles.")
    stats = None

# Importar openpyxl para Excel (opcional)
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.info("openpyxl no está instalado. La exportación a .xlsx no estará disponible.")


class AnalysisService:
    # Eliminar app_settings de la inicialización
    def __init__(self, study_service: StudyService, file_service: FileService):
        """
        Inicializa el AnalysisService.

        :param study_service: Instancia de StudyService.
        :param file_service: Instancia de FileService.
        """
        self.study_service = study_service
        self.file_service = file_service
        # self.settings = app_settings # Ya no se guarda referencia a AppSettings

    def get_analysis_parameters(self, study_id: int) -> dict:
        """
        Obtiene los parámetros disponibles para análisis de un estudio, incluyendo cálculos.

        :param study_id: ID del estudio.
        :return: Diccionario con sets de parámetros disponibles
                 {'patients': set(), 'frequencies': set(), 'descriptors_by_vi': dict, 'calculations': set()}
                 Retorna sets vacíos si no se encuentran parámetros o hay error.
        """
        try:
            # Obtener parámetros únicos del FileService (ahora devuelve 'descriptors_by_vi')
            params = self.file_service.get_unique_study_parameters(study_id)
            # Añadir cálculos fijos
            params['calculations'] = {'Maximo', 'Minimo', 'Rango'}
            # Asegurar que 'descriptors_by_vi' exista aunque esté vacío
            if 'descriptors_by_vi' not in params:
                params['descriptors_by_vi'] = {}
            return params
        except Exception as e:
            logger.error(f"Error obteniendo parámetros de análisis para estudio {study_id}: {e}", exc_info=True)
            # Devolver vacío en caso de error para que la UI no falle
            return {'patients': set(), 'frequencies': set(),
                    'descriptors': set(), 'calculations': set()}

    def _read_processed_file_data(self, file_path: Path) -> pd.DataFrame | None:
        """
        Lee los datos numéricos de un archivo procesado (.txt separado por ';').
        Omite las primeras 4 líneas de encabezado y las últimas 3 de estadísticas.
        Devuelve un DataFrame con los datos numéricos o None si hay error.
        """
        try:
            # Leer todas las líneas primero para poder omitir las últimas
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            # 4 header + 3 stats = 7. Necesita al menos 1 fila de datos.
            if len(lines) <= 7:
                logger.warning(f"Archivo {file_path.name} no contiene "
                               f"suficientes líneas para extraer datos.")
                return None

            # Omitir encabezado y estadísticas
            data_lines = lines[4:-3]

            # Usar pandas para leer los datos, especificando el separador
            # Necesitamos pasar las líneas como un stream
            from io import StringIO
            data_io = StringIO("".join(data_lines))

            # --- Determinar número de columnas de los datos reales ---
            if not data_lines:
                logger.warning(f"Archivo {file_path.name} no contiene líneas "
                               f"de datos después de quitar cabecera/pie.")
                return None
            first_data_line_parts = data_lines[0].strip().split(';')
            num_data_cols = len(first_data_line_parts)
            # Add debug for first data line content
            logger.debug(f"Primera línea de datos: '{data_lines[0].strip()}'")
            logger.debug(f"Detectadas {num_data_cols} columnas en la primera "
                         f"línea de datos de {file_path.name}")

            if num_data_cols == 0:
                logger.warning(f"No se detectaron columnas de datos en "
                               f"{file_path.name}")
                return None

            # --- Generar nombres de columna basados en línea 3 (';' sep) ---
            # Leer la línea de nombres de columna (índice 2)
            col_names_line = lines[2].strip()
            raw_col_names_from_header = col_names_line.split(';')
            logger.debug(f"Nombres crudos leídos de línea 3: {raw_col_names_from_header}")

            # Validar si el número de nombres coincide con datos
            if len(raw_col_names_from_header) != num_data_cols:
                logger.warning(f"Discrepancia en {file_path.name}: "
                               f"Nombres línea 3 ({len(raw_col_names_from_header)}) != "
                               f"Columnas datos ({num_data_cols}). "
                               f"Se usarán nombres truncados/rellenados.")
                # Podríamos fallar aquí si es grave

            # Sanear nombres de columna leídos (unicidad, no vacíos)
            final_col_names = []
            counts = {}
            # Sanitize names from the header line
            for i, name in enumerate(raw_col_names_from_header):
                clean_name = name.strip()
                # Handle empty names
                if not clean_name:
                    clean_name = f"Unnamed_{i}"  # Usar nombre genérico simple

                # Add suffix if the name is duplicated
                if clean_name in counts:
                    counts[clean_name] += 1
                    unique_name = f"{clean_name}_{counts[clean_name]}"
                else:
                    counts[clean_name] = 0
                    unique_name = clean_name
                final_col_names.append(unique_name)
            logger.debug(f"Nombres saneados ({len(final_col_names)}): {final_col_names}")

            # Ajustar lista final si hubo discrepancia con num_data_cols
            if len(final_col_names) > num_data_cols:
                final_col_names = final_col_names[:num_data_cols]  # Truncar
            elif len(final_col_names) < num_data_cols:
                # Pad con nombres genéricos
                for i in range(len(final_col_names), num_data_cols):
                    final_col_names.append(f"Data_Col_{i}")

            logger.debug(f"Nombres columna finales ajustados para "
                         f"{file_path.name} ({len(final_col_names)}): "
                         f"{final_col_names}")
            # --- Fin ajuste de nombres ---

            try:
                df = pd.read_csv(data_io, sep=';', header=None,
                                 names=final_col_names, na_values=[''],
                                 keep_default_na=True)
            except pd.errors.ParserError as pe:
                # Añadir más contexto al error de pandas
                logger.error(f"Error Pandas al parsear {file_path.name} con "
                             f"{len(final_col_names)} columnas esperadas: {pe}",
                             exc_info=True)
                raise  # Relanzar para que se maneje en el bloque exterior

            # Seleccionar solo columnas numéricas
            numeric_cols = []
            for col in df.columns:
                # Intentar convertir a numérico, ignorando 'Tiempo'
                if col.lower() == 'tiempo':
                    numeric_cols.append(col)
                    continue
                try:
                    pd.to_numeric(df[col])
                    numeric_cols.append(col)
                except (ValueError, TypeError):
                    pass  # Ignorar columnas no numéricas

            return df[numeric_cols]

        except FileNotFoundError:
            logger.error(f"Archivo no encontrado al leer datos: {file_path}")
            return None
        except Exception as e:
            logger.error(f"Error leyendo datos de {file_path.name}: {e}",
                         exc_info=True)
            return None

    def _get_data_for_parameters(self, study_id: int, parameters: dict) -> dict:
        """
        Obtiene y estructura los datos numéricos de los archivos que coinciden
        con los parámetros seleccionados.

        :param study_id: ID del estudio.
        :param parameters: Diccionario con listas de 'patients', 'frequencies', 'descriptors', 'calculations'.
        :return: Diccionario anidado:
                 {
                     'frequency1': {
                         'descriptor_combo_key': { # Clave basada en descriptores encontrados
                             'patient1': DataFrame,
                             'patient2': DataFrame, ...
                         }, ...
                     }, ...
                 Retorna diccionario vacío si no hay datos o hay error.
        """
        structured_data = {}
        # Usar método protegido para obtener ruta
        study_path = self.file_service._get_study_path(study_id)
        if not study_path:
            return {}

        selected_patients = parameters.get('patients', [])
        selected_frequencies = parameters.get('frequencies', [])
        # selected_descriptors ya no se usa directamente aquí, se usa la estructura VI

        # Obtener estructura de VIs del estudio para validación
        try:
            study_details = self.study_service.get_study_details(study_id)
            independent_variables = study_details.get('independent_variables', [])
        except Exception as e:
            logger.error(f"Error obteniendo VIs estudio {study_id}: {e}",
                         exc_info=True)
            return {}

        for patient in selected_patients:
            patient_path = study_path / patient
            if not patient_path.is_dir():
                continue

            for freq in selected_frequencies:
                freq_path = patient_path / freq
                if not freq_path.is_dir():
                    continue

                if freq not in structured_data:
                    structured_data[freq] = {}

                # Iterar sobre archivos en carpeta de frecuencia
                # Asumiendo extensión .txt para procesados
                for file_path in freq_path.glob('*.txt'):
                    filename = file_path.name

                    # Validar nombre de archivo usando VIs
                    is_valid_name, extracted_descriptors = validate_filename_for_study_criteria(
                        filename, independent_variables
                    )
                    if not is_valid_name:
                        continue # Omitir archivo si no cumple criterios

                    # Crear clave de grupo combinada basada en VIs y descriptores extraídos
                    # Formato: "VI1=DescA;VI2=DescB" o "VI1=Nulo;VI2=DescC"
                    group_parts = []
                    for i, desc in enumerate(extracted_descriptors):
                        vi_name = independent_variables[i].get('name', f'VI{i+1}') # Usar nombre VI
                        value = desc if desc is not None else "Nulo" # Usar "Nulo" si es None
                        group_parts.append(f"{vi_name}={value}")

                    # Usar ';' como separador para evitar conflictos con nombres
                    group_key = ";".join(group_parts) if group_parts else "SinGrupo"

                    if group_key not in structured_data[freq]:
                        structured_data[freq][group_key] = {}

                    # Leer datos del archivo
                    # (La lógica de lectura y concatenación permanece igual)
                    df_data = self._read_processed_file_data(file_path)
                    if df_data is not None and not df_data.empty:
                        # Acumular datos si ya existe una entrada para este paciente/freq/group_key
                        if patient not in structured_data[freq][group_key]:
                            structured_data[freq][group_key][patient] = df_data
                        else:
                            # Concatenar DataFrames
                            structured_data[freq][group_key][patient] = \
                                pd.concat(
                                    [structured_data[freq][group_key][patient], df_data],
                                    ignore_index=True
                                )
                    else:
                        logger.warning(f"No se pudieron leer datos válidos de {filename}")

        return structured_data

    def _calculate_statistic(self, df: pd.DataFrame, calculation: str) -> pd.Series | None:
        """Calcula una estadística ('Maximo', 'Minimo', 'Rango') para cada columna numérica del DataFrame."""
        if df is None or df.empty:
            return None

        # Seleccionar solo columnas numéricas (excluir 'Tiempo')
        numeric_df = df.select_dtypes(include=np.number)
        if 'Tiempo' in numeric_df.columns:
            numeric_df = numeric_df.drop(columns=['Tiempo'])

        # Devolver None si no quedan columnas numéricas O si todas son NaN
        if numeric_df.empty or numeric_df.isnull().all().all():
            return None

        if calculation == "Maximo":
            return numeric_df.max(skipna=True)
        elif calculation == "Minimo":
            return numeric_df.min(skipna=True)
        elif calculation == "Rango":
            return numeric_df.max(skipna=True) - numeric_df.min(skipna=True)
        else:
            logger.warning(f"Cálculo no soportado '{calculation}'")
            return None

    def perform_analysis(self, study_id: int, parameters: dict):
        """
        Realiza un análisis basado en los parámetros proporcionados.
        Calcula las estadísticas seleccionadas para los datos agrupados.

        :param study_id: ID del estudio a analizar.
        :param parameters: Diccionario con los parámetros de análisis ('patients', 'frequencies', 'descriptors', 'calculations').
                 frecuencia -> descriptor_key -> calculo -> paciente -> Serie.
                 Ej: {'Cinematica': {'CMJ_PRE': {'Maximo': {'P01': pd.Series}}}}
        """
        logger.info(f"Realizando análisis para estudio {study_id} con "
                    f"parámetros: {parameters}")
        structured_data = self._get_data_for_parameters(study_id, parameters)
        analysis_results = {}
        selected_calculations = parameters.get('calculations', [])

        if not structured_data:
            logger.warning(f"No se encontraron datos para los parámetros "
                           f"seleccionados en estudio {study_id}.")
            return {}

        for freq, group_data in structured_data.items(): # Cambiar descriptor_data a group_data
            analysis_results[freq] = {}
            for group_key, patient_data in group_data.items(): # Cambiar descriptor_key a group_key
                analysis_results[freq][group_key] = {}
                for calc in selected_calculations:
                    analysis_results[freq][group_key][calc] = {}
                    for patient, df in patient_data.items():
                        stats = self._calculate_statistic(df, calc)
                        if stats is not None:
                            analysis_results[freq][group_key][calc][patient] = stats

        logger.info(f"Análisis completado para estudio {study_id}.")
        return analysis_results


    def generate_report(self, study_id: int, parameters: dict, output_path_str: str):
        """
        Genera un reporte PDF del análisis.

        :param study_id: ID del estudio.
        :param parameters: Parámetros del análisis.
        :param output_path_str: Ruta (string) donde guardar el reporte PDF.
        """
        logger.info(f"Generando reporte para estudio {study_id} en "
                    f"{output_path_str}...")
        output_path = Path(output_path_str)
        # Asegurar que directorio exista
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # --- Obtener Datos y Detalles ---
        try:
            study_details = self.study_service.get_study_details(study_id)
            study_name = study_details.get('name', f'Estudio {study_id}')
        except Exception as e:
            raise ValueError(f"No se pudieron obtener detalles del estudio "
                             f"{study_id}: {e}")

        # Obtener datos estructurados (agrupados por paciente)
        structured_data = self._get_data_for_parameters(study_id, parameters)
        selected_calculations = parameters.get('calculations', [])
        selected_patients = parameters.get('patients', [])

        if not structured_data:
            raise ValueError("No se encontraron datos para generar el reporte "
                             "con los parámetros seleccionados.")

        # --- Crear Directorio Temporal para Gráficos ---
        # Usar tempfile para limpieza automática si falla
        with tempfile.TemporaryDirectory(
                prefix=f"kineviz_report_{study_id}_") as temp_dir_str:
            temp_dir = Path(temp_dir_str)
            logger.debug(f"Directorio temporal para gráficos: {temp_dir}")

            # --- Configurar PDF con ReportLab ---
            doc = SimpleDocTemplate(
                output_path_str, pagesize=letter,
                leftMargin=0.75*inch, rightMargin=0.75*inch,
                topMargin=1*inch, bottomMargin=1*inch
            )
            styles = getSampleStyleSheet()
            story = []

            # --- Título y Metadatos ---
            story.append(Paragraph(f"Reporte de Análisis - {study_name}",
                                   styles['h1']))
            story.append(Spacer(1, 0.2*inch))
            story.append(Paragraph(f"Fecha de Generación: "
                                   f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                   styles['Normal']))
            story.append(Spacer(1, 0.1*inch))

            # Parámetros Usados
            story.append(Paragraph("<b>Parámetros Seleccionados:</b>",
                                   styles['h3']))
            # Obtener alias del estudio
            study_aliases = self.study_service.get_study_aliases(study_id)
            # Mostrar parámetros seleccionados (ya no incluye 'descriptors' directamente)
            param_text = (
                f"<b>Pacientes:</b> {', '.join(parameters.get('patients', []))}<br/>"
                f"<b>Frecuencias:</b> {', '.join(parameters.get('frequencies', []))}<br/>"
                # Podríamos añadir VIs/Descriptores si se seleccionaron explícitamente,
                # pero por ahora omitimos esa parte ya que el análisis agrupa por todas las combinaciones.
                f"<b>Cálculos:</b> {', '.join(parameters.get('calculations', []))}"
            )
            story.append(Paragraph(param_text, styles['BodyText']))
            story.append(Spacer(1, 0.3*inch))

            # --- Iterar y Generar Contenido ---
            plot_counter = 0
            for freq, descriptor_data in structured_data.items():
                story.append(Paragraph(f"Resultados para Frecuencia: {freq}",
                                       styles['h2']))
                story.append(Spacer(1, 0.1*inch))

                for group_key, patient_data in descriptor_data.items(): # Usar descriptor_data
                    # Convertir group_key a formato legible con alias
                    group_display_parts = []
                    if group_key != "SinGrupo":
                        for part in group_key.split(';'):
                            vi_name, desc_value = part.split('=', 1)
                            alias = study_aliases.get(desc_value, desc_value) # Aplicar alias
                            group_display_parts.append(f"{vi_name}: {alias}")
                    group_display = ", ".join(group_display_parts) if group_display_parts else "Grupo General"

                    story.append(Paragraph(f"Grupo: {group_display}", styles['h3']))
                    story.append(Spacer(1, 0.1*inch))

                    # --- Boxplot General por Paciente ---
                    boxplot_data = {}
                    all_numeric_columns = set()
                    for patient, df in patient_data.items():
                        numeric_df = df.select_dtypes(include=np.number)
                        if 'Tiempo' in numeric_df.columns:
                            numeric_df = numeric_df.drop(columns=['Tiempo'])
                        if not numeric_df.empty:
                            # Usar todos los valores de todas las columnas
                            boxplot_data[patient] = numeric_df.values.flatten()
                            all_numeric_columns.update(numeric_df.columns)

                    if boxplot_data:
                        plot_counter += 1
                        boxplot_filename = temp_dir / f"boxplot_{plot_counter}.png"
                        # Usar group_display (con alias) en el título del gráfico
                        charting.create_boxplot(
                            data_dict=boxplot_data,
                            title=f"Distribución General - {freq} ({group_display})",
                            ylabel="Valor Medición",
                            output_path=boxplot_filename
                        )
                        if boxplot_filename.exists():
                            # Ajustar tamaño
                            story.append(Image(str(boxplot_filename),
                                               width=6*inch, height=4*inch))
                            story.append(Spacer(1, 0.2*inch))
                        else:
                            story.append(Paragraph(f"<i>Error generando boxplot "
                                                   f"{plot_counter}</i>",
                                                   styles['Italic']))
                    else:
                        story.append(Paragraph("<i>No hay datos suficientes "
                                               "para el boxplot general.</i>",
                                               styles['Italic']))

                    # --- Cálculos y Gráficos de Barras ---
                    for calc in selected_calculations:
                        story.append(Paragraph(f"<b>Cálculo: {calc}</b>",
                                               styles['Normal']))
                        story.append(Spacer(1, 0.05*inch))

                        calc_results_by_patient = {}
                        valid_columns_for_calc = set()
                        for patient, df in patient_data.items():
                            stats = self._calculate_statistic(df, calc)
                            if stats is not None and not stats.empty:
                                calc_results_by_patient[patient] = stats
                                valid_columns_for_calc.update(stats.index)

                        if calc_results_by_patient:
                            # --- Gráfico de Barras (Promedio por Paciente) ---
                            # Calcular promedio del cálculo para cada paciente
                            avg_calc_per_patient = {
                                patient: results.mean(skipna=True)
                                for patient, results
                                in calc_results_by_patient.items()
                                if results is not None
                            }
                            # Filtrar pacientes sin promedio válido
                            valid_avg_calc = {p: v for p, v
                                              in avg_calc_per_patient.items()
                                              if pd.notna(v)}

                            if valid_avg_calc:
                                plot_counter += 1
                                barchart_filename = temp_dir / f"barchart_{plot_counter}.png"
                                # Usar group_display (con alias) en el título del gráfico
                                charting.create_barchart(
                                    data_dict=valid_avg_calc,
                                    title=f"{calc} Promedio - {freq} ({group_display})",
                                    xlabel="Paciente",
                                    ylabel=f"{calc} Promedio",
                                    output_path=barchart_filename
                                )
                                if barchart_filename.exists():
                                    story.append(Image(str(barchart_filename),
                                                       width=6*inch, height=4*inch))
                                    story.append(Spacer(1, 0.1*inch))
                                else:
                                    story.append(Paragraph(f"<i>Error generando "
                                                           f"barchart {plot_counter}</i>",
                                                           styles['Italic']))
                            else:
                                story.append(Paragraph(f"<i>No hay datos suficientes "
                                                       f"para el gráfico de barras "
                                                       f"de {calc}.</i>",
                                                       styles['Italic']))

                            # --- Tabla de Resultados Detallados ---
                            # Crear tabla: pacientes (filas) x mediciones (cols)
                            if valid_columns_for_calc:
                                sorted_columns = sorted(list(valid_columns_for_calc))
                                table_data = [['Paciente'] + sorted_columns]  # Cabecera
                                # Iterar en orden de selección
                                for patient in selected_patients:
                                    if patient in calc_results_by_patient:
                                        results = calc_results_by_patient[patient]
                                        row_values = [
                                            f"{results.get(col, 'N/A'):.3f}"
                                            if pd.notna(results.get(col)) else 'N/A'
                                            for col in sorted_columns
                                        ]
                                        row = [patient] + row_values
                                        table_data.append(row)
                                    # else: # Opcional: fila si no hay datos
                                    #     table_data.append([patient] + ['N/A'] * len(sorted_columns))

                                # Si hay datos además de la cabecera
                                if len(table_data) > 1:
                                    table = Table(table_data, hAlign='LEFT')
                                    table.setStyle(TableStyle([
                                        ('BACKGROUND', (0, 0), (-1, 0),
                                         colors.grey),
                                        ('TEXTCOLOR', (0, 0), (-1, 0),
                                         colors.whitesmoke),
                                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                        ('FONTNAME', (0, 0), (-1, 0),
                                         'Helvetica-Bold'),
                                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                        ('BACKGROUND', (0, 1), (-1, -1),
                                         colors.beige),
                                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                                    ]))
                                    story.append(table)
                                    story.append(Spacer(1, 0.2*inch))
                                else:
                                    story.append(Paragraph(f"<i>No hay resultados "
                                                           f"detallados para {calc}.</i>",
                                                           styles['Italic']))
                            else:
                                story.append(Paragraph(f"<i>No hay columnas válidas "
                                                       f"para mostrar resultados "
                                                       f"detallados de {calc}.</i>",
                                                       styles['Italic']))
                        else:
                            story.append(Paragraph(f"<i>No se pudieron calcular "
                                                   f"resultados para {calc}.</i>",
                                                   styles['Italic']))
                        # Espacio después de cada cálculo
                        story.append(Spacer(1, 0.1*inch))

                    # Espacio después de cada tipo/periodo
                    story.append(Spacer(1, 0.2*inch))

            # --- Construir PDF ---
            try:
                doc.build(story)
                logger.info(f"Reporte PDF generado exitosamente en {output_path}")
            except Exception as build_e:
                logger.error(f"Error construyendo el PDF para estudio "
                             f"{study_id}: {build_e}", exc_info=True)
                raise  # Relanzar error de construcción

        # El directorio temporal se limpia automáticamente al salir del 'with'

    # --- Métodos para Gestión de Reportes ---

    def list_reports(self, study_id: int) -> list[dict]:
        """
        Lista los archivos PDF de reportes generados para un estudio.

        :param study_id: ID del estudio.
        :return: Lista de diccionarios {'name': str, 'path': str}.
                 Retorna lista vacía si no hay reportes o hay error.
        """
        reports = []
        try:
            study_path = self.file_service._get_study_path(study_id)
            if not study_path:
                return []

            reports_dir = study_path / "reportes"
            if reports_dir.exists() and reports_dir.is_dir():
                for item in reports_dir.glob("*.pdf"):
                    if item.is_file():
                        reports.append({'name': item.name, 'path': str(item)})
                # Ordenar por nombre (o fecha si se extrae)
                reports.sort(key=lambda x: x['name'], reverse=True)
        except Exception as e:
            logger.error(f"Error listando reportes para estudio {study_id}: {e}",
                         exc_info=True)
        return reports

    def delete_report(self, report_path_str: str):
        """
        Elimina un archivo de reporte específico.

        :param report_path_str: Ruta completa (string) del archivo PDF a eliminar.
        :raises FileNotFoundError: Si el archivo no existe.
        :raises OSError: Si ocurre un error al eliminar.
        """
        report_path = Path(report_path_str)
        if not report_path.exists():
            raise FileNotFoundError(f"El archivo de reporte no existe: "
                                    f"{report_path}")
        if not report_path.is_file():
            raise ValueError(f"La ruta no es un archivo: {report_path}")

        try:
            report_path.unlink()
            logger.info(f"Reporte eliminado: {report_path}")
        except OSError as e:
            logger.error(f"Error al eliminar el reporte {report_path}: {e}",
                         exc_info=True)
            raise

    # --- Métodos para Análisis Discreto (Fase 6) ---

    def get_discrete_analysis_tables_path(self, study_id: int) -> Path | None:
        """
        Obtiene la ruta base donde se guardan las tablas de resumen discreto.

        :param study_id: ID del estudio.
        :return: Path al directorio de tablas o None si no se puede determinar.
        """
        try:
            study_path = self.file_service._get_study_path(study_id)
            if not study_path:
                logger.warning(f"No se pudo obtener ruta estudio {study_id} "
                               f"para buscar tablas discretas.")
                return None
            # Asume estructura definida en generate_discrete_summary_tables
            # Nota: Podría ser más robusto guardando ruta relativa en config.
            tables_path = study_path / "Analisis Discreto" / "Tablas"
            return tables_path
        except Exception as e:
            logger.error(f"Error obteniendo ruta tablas discretas para "
                         f"estudio {study_id}: {e}", exc_info=True)
            return None

    def delete_discrete_summary_table(self, table_path_str: str):
        """
        Elimina un archivo de tabla de resumen discreto específico.

        :param table_path_str: Ruta completa (string) del archivo CSV a eliminar.
        :raises FileNotFoundError: Si el archivo no existe.
        :raises ValueError: Si la ruta no es un archivo.
        :raises OSError: Si ocurre un error al eliminar.
        """
        table_path = Path(table_path_str)
        if not table_path.exists():
            raise FileNotFoundError(f"El archivo de tabla no existe: "
                                    f"{table_path}")
        if not table_path.is_file():
            raise ValueError(f"La ruta no es un archivo: {table_path}")
        if not table_path.name.endswith('.csv'):
            raise ValueError(f"El archivo no parece ser una tabla CSV: "
                             f"{table_path}")

        try:
            table_path.unlink()
            logger.info(f"Tabla de resumen discreto eliminada: {table_path}")
            # Opcional: Limpiar directorios vacíos (Frecuencia, etc.)
            # Requeriría lógica adicional para verificar si están vacíos.
        except OSError as e:
            logger.error(f"Error al eliminar la tabla {table_path}: {e}",
                         exc_info=True)
            raise

    def _extract_stats_from_processed_file(self, file_path: Path, calculation: str) -> list | None:
        """Lee las últimas líneas de un archivo procesado y extrae la fila de datos para el cálculo especificado."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            if len(lines) < 3:  # Necesita al menos las 3 líneas de stats
                logger.warning(f"Archivo {file_path.name} no tiene suficientes "
                               f"líneas para extraer estadísticas.")
                return None

            # Buscar la línea del cálculo en las últimas 3 líneas
            calc_line_prefix = f";;{calculation.upper()};"
            for line in reversed(lines[-3:]):
                if line.startswith(calc_line_prefix):
                    # Quitar prefijo y dividir por ';'
                    # Devolver valores como lista de strings (incluye vacíos)
                    return line.strip()[len(calc_line_prefix):].split(';')
            logger.warning(f"No se encontró línea de cálculo '{calculation}' "
                           f"en {file_path.name}")
            return None
        except Exception as e:
            logger.error(f"Error extrayendo stats de {file_path.name} para "
                         f"cálculo {calculation}: {e}", exc_info=True)
            return None

    def _parse_processed_file_headers(self, file_path: Path) -> tuple[list, list, list] | None:
        """Lee las líneas 1, 2 y 3 (atributos, columnas, unidades) de un archivo procesado."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            if len(lines) < 4:  # Necesita num_frames, attr, col, unit
                logger.warning(f"Archivo {file_path.name} no tiene suficientes "
                               f"líneas de cabecera.")
                return None
            # Líneas 1, 2, 3 (índices 1, 2, 3)
            # Omitir Frame, SubFrame, Tiempo y limpiar prefijo PteXX:
            atributos_raw = lines[1].strip().split(';')[3:]
            columnas_raw = lines[2].strip().split(';')[3:]
            unidades_raw = lines[3].strip().split(';')[3:]

            # Función auxiliar para limpiar prefijo "PteXX:"
            def clean_prefix(name: str) -> str:
                name = name.strip()
                if ':' in name:
                    # Intentar dividir y tomar la segunda parte
                    parts = name.split(':', 1)
                    if len(parts) > 1:
                        return parts[1].strip()
                # Si no hay ':' o la división falla, devolver el nombre original (limpio)
                return name

            # Limpiar prefijos de atributos y columnas
            atributos = [clean_prefix(a) for a in atributos_raw]
            columnas = [clean_prefix(c) for c in columnas_raw]
            # Limpiar espacios en unidades (no deberían tener prefijo)
            unidades = [u.strip() for u in unidades_raw]

            # Asegurar que todas las listas tengan la misma longitud
            max_len = max(len(atributos), len(columnas), len(unidades))
            atributos.extend([''] * (max_len - len(atributos)))
            columnas.extend([''] * (max_len - len(columnas)))
            unidades.extend([''] * (max_len - len(unidades)))

            return atributos, columnas, unidades
        except Exception as e:
            logger.error(f"Error parseando cabeceras de {file_path.name}: {e}",
                         exc_info=True)
            return None

    def generate_discrete_summary_tables(self, study_id: int):
        """
        Genera tablas resumen CSV para cálculos discretos (Max, Min, Rango)
        agrupados por frecuencia y combinación de descriptores.
        Enfocado inicialmente en 'Cinematica'.

        :param study_id: ID del estudio.
        :return: Diccionario con rutas de los archivos generados o errores.
                 {'success': [path_str], 'errors': [error_msg]}
        """
        logger.info(f"Iniciando generación de tablas resumen discreto para "
                    f"estudio {study_id}")
        results = {'success': [], 'errors': []}
        target_frequency = "Cinematica"  # Enfocarse en Cinemática por ahora
        calculations = ["Maximo", "Minimo", "Rango"]

        try:
            study_path = self.file_service._get_study_path(study_id)
            if not study_path:
                results['errors'].append(f"No se pudo encontrar la ruta del estudio {study_id}.")
                return results

            study_details = self.study_service.get_study_details(study_id)
            defined_descriptors = [d.strip() for d in
                                   (study_details.get('descriptores', '') or '')
                                   .split(',') if d.strip()]

            # 1. Encontrar y agrupar archivos procesados de Cinemática
            # { 'Desc1_Desc2': [path1, path2,...], ... }
            files_by_descriptor_combo = {}
            processed_files_info, _ = self.file_service.get_study_files( # Corregir nombre variable
                study_id=study_id,
                page=1, # Eliminar el 'page=1' duplicado
                per_page=10000, file_type='Processed', frequency=target_frequency
            )
            path_map = {f_info['path'].stem.split(f'_{target_frequency}')[0]: f_info['path']
                        for f_info in processed_files_info}

            # --- Identificar grupos y mapear archivos ---
            # Usar _identify_study_groups para obtener el mapeo archivo_base -> group_key
            try:
                files_to_groups, _ = self._identify_study_groups(study_id, target_frequency)
            except ValueError as e_group:
                results['errors'].append(f"Error identificando grupos para {target_frequency}: {e_group}")
                return results

            files_by_group_key = {} # Inicializar aquí
            for file_base_key, group_key in files_to_groups.items():
                if group_key not in files_by_group_key:
                    files_by_group_key[group_key] = []
                # Buscar la ruta completa usando el file_base_key
                full_path = path_map.get(file_base_key)
                if full_path:
                    files_by_group_key[group_key].append(full_path)
                else:
                    logger.warning(f"No se encontró la ruta completa para el archivo base '{file_base_key}' del grupo '{group_key}'.")


            if not files_by_group_key:
                results['errors'].append(f"No se encontraron archivos válidos agrupables por VIs para '{target_frequency}'.")
                return results

            # 2. Preparar directorio de salida (limpiar si existe)
            output_base_dir = (study_path / "Analisis Discreto" / "Tablas" /
                               target_frequency)
            if output_base_dir.exists():
                logger.info(f"Limpiando directorio de salida existente: {output_base_dir}")
                try:
                    for item in output_base_dir.iterdir():
                        if item.is_dir():
                            shutil.rmtree(item)
                        else:
                            item.unlink()
                except OSError as e:
                    logger.error(f"Error limpiando directorio {output_base_dir}: {e}", exc_info=True)
                    # Continuar de todos modos, puede que solo fallen algunos archivos
            output_base_dir.mkdir(parents=True, exist_ok=True)

            # 3. Generar tabla para cada grupo y cálculo
            for group_key, file_paths in files_by_group_key.items():
                if not file_paths:
                    continue

                # Crear nombre de archivo seguro reemplazando caracteres inválidos
                safe_group_key_part = group_key.replace('=', '_').replace(';', '__')

                # Leer cabeceras desde el primer archivo del grupo
                headers = self._parse_processed_file_headers(file_paths[0])
                if not headers:
                    results['errors'].append(f"No se pudieron leer las cabeceras para el grupo '{group_key}'.")
                    continue
                atributos, columnas, unidades = headers
                # Número de columnas de datos (sin Frame, Sub, Tiempo)
                num_value_cols = len(columnas)

                # Crear MultiIndex para las columnas
                multi_index_tuples = []
                last_attr = ""
                for i in range(num_value_cols):
                    # Propagar atributo si está vacío
                    attr = atributos[i] if atributos[i] else last_attr
                    # Añadir unidad
                    multi_index_tuples.append((attr, columnas[i], unidades[i]))
                    last_attr = attr
                # Crear MultiIndex con tres niveles
                column_multi_index = pd.MultiIndex.from_tuples(
                    multi_index_tuples,
                    names=["Atributo", "Columna", "Unidad"]
                )


                for calc in calculations:
                    table_data = []
                    # Almacenar nombres base completos para el índice
                    index_names = []

                    for file_path in file_paths:
                        stats_row = self._extract_stats_from_processed_file(
                            file_path, calc
                        )
                        # Obtener nombre base completo para índice
                        base_name = file_path.stem.split(f'_{target_frequency}')[0]

                        if stats_row and len(stats_row) == num_value_cols:
                            table_data.append(stats_row)
                            index_names.append(base_name) # Guardar nombre base completo
                        else:
                            logger.warning(f"Datos de '{calc}' inconsistentes o "
                                           f"faltantes en {file_path.name}. Se "
                                           f"omitirá de las tablas resumen.")
                            # No añadir a index_names si se omite la fila

                    if table_data:
                        try:
                            # --- Crear y Guardar CSV Interno (para lógica posterior) ---
                            df_csv_internal = pd.DataFrame(table_data, columns=column_multi_index, index=index_names)
                            df_csv_internal.index.name = "ARCHIVO"

                            # Convertir CSV interno a numérico (punto decimal)
                            for col in df_csv_internal.columns:
                                if df_csv_internal[col].dtype == 'object':
                                    df_csv_internal[col] = df_csv_internal[col].str.replace(',', '.', regex=False)
                                df_csv_internal[col] = pd.to_numeric(df_csv_internal[col], errors='coerce')

                            # Usar safe_group_key_part para el nombre de archivo
                            output_csv_internal_path = output_base_dir / f"{calc}_{target_frequency}_{safe_group_key_part}.csv"
                            df_csv_internal.to_csv(output_csv_internal_path, sep=',', decimal='.',
                                                   encoding='utf-8', header=True, index=True)
                            results['success'].append(str(output_csv_internal_path))
                            logger.info(f"Tabla CSV interna generada: {output_csv_internal_path}")

                            # --- Preparar DataFrame para formatos de exportación (TSV, XLSX, SCSV) ---
                            # Usar los mismos datos numéricos pero con índice como columna 'ARCHIVO'
                            df_data_export = df_csv_internal.reset_index() # Mover índice 'ARCHIVO' a columna

                            # --- Construir Cabeceras Manualmente (para TSV, XLSX, SCSV) ---
                            # 4 filas de cabecera según el ejemplo TSV
                            header_rows_list = []
                            num_data_cols = len(column_multi_index)
                            total_cols = 1 + num_data_cols # ARCHIVO + datos

                            # Fila 0: Vacía (solo separadores)
                            header_rows_list.append([''] * total_cols)

                            # Fila 1: Caracteristica (Atributo) - Empieza vacía
                            row1 = [''] # Celda vacía para columna ARCHIVO
                            last_attr = None
                            for attr, _, _ in column_multi_index:
                                if attr == last_attr and last_attr is not None:
                                    row1.append('') # Simular celda combinada
                                else:
                                    row1.append(attr if attr else '')
                                    last_attr = attr
                            header_rows_list.append(row1)

                            # Fila 2: Espacio (Columna) - Empieza con 'ARCHIVO'
                            row2 = ['ARCHIVO']
                            for _, col, _ in column_multi_index:
                                row2.append(col if col else '')
                            header_rows_list.append(row2)

                            # Fila 3: Unidad - Empieza vacía
                            row3 = [''] # Celda vacía para columna ARCHIVO
                            for _, _, unit in column_multi_index:
                                row3.append(unit if unit else '')
                            header_rows_list.append(row3)

                            # --- Guardar Formatos de Exportación (usando safe_group_key_part) ---

                            # 1. Formato TSV (Tab Separated, Comma Decimal)
                            output_tsv_path = output_base_dir / f"{calc}_{target_frequency}_{safe_group_key_part}.tsv"
                            try:
                                with open(output_tsv_path, 'w', encoding='utf-8') as f:
                                    # Escribir las 4 filas de cabecera
                                    for header_row_list in header_rows_list:
                                        f.write('\t'.join(header_row_list) + '\n')
                                    # Escribir datos, coma decimal
                                    df_data_export.to_csv(f, sep='\t', decimal=',', header=False, index=False,
                                                          encoding='utf-8', float_format='%.4f')
                                results['success'].append(str(output_tsv_path))
                                logger.info(f"Tabla TSV generada: {output_tsv_path}")
                            except Exception as e_tsv:
                                error_msg = f"Error guardando TSV {output_tsv_path.name}: {e_tsv}"
                                logger.error(error_msg, exc_info=True)
                                results['errors'].append(error_msg)

                            # 2. Formato SCSV (Semicolon Separated, Comma Decimal)
                            output_scsv_path = output_base_dir / f"{calc}_{target_frequency}_{safe_group_key_part}.scsv"
                            try:
                                with open(output_scsv_path, 'w', encoding='utf-8') as f:
                                    # Escribir las 4 filas de cabecera
                                    for header_row_list in header_rows_list:
                                        f.write(';'.join(header_row_list) + '\n')
                                    # Escribir datos, coma decimal
                                    df_data_export.to_csv(f, sep=';', decimal=',', header=False, index=False,
                                                          encoding='utf-8', float_format='%.4f')
                                results['success'].append(str(output_scsv_path))
                                logger.info(f"Tabla SCSV generada: {output_scsv_path}")
                            except Exception as e_scsv:
                                error_msg = f"Error guardando SCSV {output_scsv_path.name}: {e_scsv}"
                                logger.error(error_msg, exc_info=True)
                                results['errors'].append(error_msg)

                            # 3. Formato XLSX (Excel, Comma Decimal via Locale?) - openpyxl requerido
                            if OPENPYXL_AVAILABLE:
                                output_xlsx_path = output_base_dir / f"{calc}_{target_frequency}_{safe_group_key_part}.xlsx"
                                try:
                                    # Usar ExcelWriter para acceder al objeto worksheet
                                    with pd.ExcelWriter(output_xlsx_path, engine='openpyxl') as writer:
                                        # Asegurar que la hoja 'Data' exista
                                        # Si el libro está vacío, crea una hoja activa
                                        if not writer.book.sheetnames:
                                            ws = writer.book.create_sheet(title='Data')
                                        else:
                                            # Si ya hay hojas, intenta obtener 'Data' o la activa
                                            ws = writer.book.get_sheet_by_name('Data') \
                                                 if 'Data' in writer.book.sheetnames \
                                                 else writer.book.active
                                            ws.title = 'Data' # Asegurar nombre

                                        # Escribir las 4 filas de cabecera manualmente
                                        for header_row_list in header_rows_list:
                                            # openpyxl es 1-based index
                                            ws.append(header_row_list)

                                        # --- Aplanar columnas para evitar error de pandas ---
                                        # Crear copia para no modificar el original usado en TSV/SCSV
                                        df_data_export_flat = df_data_export.copy()
                                        # Crear nombres simples: 'ARCHIVO' y 'Attr/Col/Unit'
                                        flat_column_names = [df_data_export_flat.columns[0]] # Mantener 'ARCHIVO'
                                        flat_column_names.extend([
                                            '/'.join(map(str, col_tuple))
                                            for col_tuple in df_data_export_flat.columns[1:] # Iterar sobre el MultiIndex
                                        ])
                                        df_data_export_flat.columns = flat_column_names

                                        # --- Escribir datos con columnas aplanadas ---
                                        # Empezar desde la fila después de las cabeceras
                                        df_data_export_flat.to_excel(
                                            writer,
                                            sheet_name='Data',
                                            startrow=len(header_rows_list), # Fila donde empiezan los datos (0-based)
                                            header=False, # No escribir cabecera de pandas
                                            index=False   # No escribir índice de pandas
                                        )
                                    results['success'].append(str(output_xlsx_path))
                                    logger.info(f"Tabla XLSX generada: {output_xlsx_path}")
                                except Exception as e_xlsx:
                                    # Capturar el error específico si aún ocurre, o errores generales
                                    error_msg = f"Error guardando XLSX {output_xlsx_path.name}: {e_xlsx}"
                                    logger.error(error_msg, exc_info=True)
                                    results['errors'].append(error_msg)
                            else:
                                logger.warning(f"Omitiendo generación XLSX para {group_key}/{calc} (openpyxl no disponible).")


                        except Exception as e_df:
                            # Error general al procesar este cálculo/grupo
                            error_msg = (f"Error procesando datos para "
                                         f"{calc}_{target_frequency}_{group_key}: {e_df}")
                            logger.error(error_msg, exc_info=True)
                            results['errors'].append(error_msg)
                    else:
                        # Mensaje genérico ya que se generan múltiples formatos
                        logger.warning(f"No se encontraron datos válidos para "
                                       f"generar tablas resumen para "
                                       f"{calc}_{target_frequency}_{group_key}")

        except Exception as e:
            error_msg = (f"Error inesperado durante generación tablas discretas "
                         f"para estudio {study_id}: {e}")
            logger.critical(error_msg, exc_info=True)
            results['errors'].append(error_msg)

        logger.info(f"Generación tablas discretas finalizada para estudio "
                    f"{study_id}. Éxitos: {len(results['success'])}, "
                    f"Errores: {len(results['errors'])}")
        return results

    # --- Métodos para Análisis Individual (Fase 6) ---

    def _identify_study_groups(self, study_id: int, frequency: str = "Cinematica") -> tuple[dict[str, str], set[str]]:
        """
        Identifica los grupos únicos basados en descriptores de archivos procesados.

        :param study_id: ID del estudio.
        :param frequency: Frecuencia a considerar (por defecto 'Cinematica').
        :return: Tupla:
                 - Dict mapeando nombre base archivo a clave grupo (formato VI=Desc).
                 - Set de claves de grupo únicas encontradas.
        :raises ValueError: Si no se pueden obtener detalles o archivos.
        """
        logger.debug(f"Identificando grupos para estudio {study_id}, frecuencia {frequency}")
        groups_by_file_base = {}
        unique_group_keys = set()

        try:
            study_details = self.study_service.get_study_details(study_id)
            if not study_details:
                raise ValueError(f"No se pudieron obtener detalles del estudio {study_id}")
            # Obtener estructura VI
            independent_variables = study_details.get('independent_variables', [])

            processed_files, _ = self.file_service.get_study_files(
                study_id=study_id,
                page=1,
                per_page=10000,  # Obtener todos
                file_type='Processed',
                frequency=frequency
            )

            if not processed_files:
                logger.warning(f"No se encontraron archivos procesados de "
                               f"'{frequency}' para identificar grupos en "
                               f"estudio {study_id}.")
                return {}, set()

            for file_info in processed_files:
                file_path = file_info['path']
                filename = file_path.name

                # Validar nombre usando VIs
                is_valid_name, extracted_descriptors = validate_filename_for_study_criteria(
                    filename, independent_variables
                )
                if not is_valid_name:
                    continue

                # Crear clave de grupo combinada (VI=Desc;...)
                try:
                    group_parts = []
                    for i, desc in enumerate(extracted_descriptors):
                        vi_name = independent_variables[i].get('name', f'VI{i+1}')
                        value = desc if desc is not None else "Nulo"
                        group_parts.append(f"{vi_name}={value}")
                    group_key = ";".join(group_parts) if group_parts else "SinGrupo"

                    # Usar nombre base sin frecuencia ni extensión como clave del dict
                    file_base_key = file_path.stem.split(f'_{frequency}')[0]

                    groups_by_file_base[file_base_key] = group_key
                    unique_group_keys.add(group_key)
                except IndexError: # Si extracted_descriptors no coincide con VIs
                    logger.warning(f"Discrepancia VIs/Descriptores al extraer grupo de: {filename}")
                    continue

            logger.debug(f"Grupos identificados ({len(unique_group_keys)}): "
                         f"{unique_group_keys}")
            return groups_by_file_base, unique_group_keys

        except Exception as e:
            logger.error(f"Error identificando grupos para estudio {study_id}: {e}", exc_info=True)
            raise ValueError(f"Error identificando grupos: {e}")


    def get_discrete_analysis_groups(self, study_id: int, frequency: str = "Cinematica") -> list[str]:
        """
        Obtiene la lista de claves de grupos únicos para análisis discreto.

        :param study_id: ID del estudio.
        :param frequency: Frecuencia a considerar.
        :return: Lista ordenada de tuplas (display_name, original_key).
        """
        try:
            _, unique_group_keys = self._identify_study_groups(study_id, frequency)
            study_aliases = self.study_service.get_study_aliases(study_id)

            # Crear tuplas ("Grupo X - Display Name", original_key)
            groups_with_display_names = []
            # Ordenar claves originales para numeración consistente
            sorted_group_keys = sorted(list(unique_group_keys))
            for i, group_key in enumerate(sorted_group_keys):
                display_parts = []
                if group_key != "SinGrupo":
                    for part in group_key.split(';'):
                        vi_name, desc_value = part.split('=', 1)
                        alias = study_aliases.get(desc_value, desc_value) # Aplicar alias
                        display_parts.append(f"{vi_name}: {alias}") # Formato "VI: Alias"
                base_display_name = ", ".join(display_parts) if display_parts else "Grupo General"
                # Añadir prefijo "Grupo X - "
                full_display_name = f"Grupo {i+1} - {base_display_name}"
                groups_with_display_names.append((full_display_name, group_key)) # Guardar clave original

            # Ordenar por el nombre completo visible (ya incluye número)
            groups_with_display_names.sort()
            return groups_with_display_names

        except ValueError as e:
            logger.warning(f"No se pudieron obtener grupos para estudio {study_id}: {e}")
            return [] # Devolver vacío si hay error


    def get_filtered_discrete_analysis_groups(self, study_id: int, frequency: str, mode: str,
                                              primary_vi_name: Optional[str] = None,
                                              fixed_vi_name: Optional[str] = None,
                                              fixed_descriptor_value: Optional[str] = None) -> dict[str, str]:
        """
        Obtiene los grupos de análisis discreto FILTRADOS según el modo y selecciones.

        :param study_id: ID del estudio.
        :param frequency: Frecuencia seleccionada.
        :param mode: '1VI' o '2VIs'.
        :param primary_vi_name: Nombre de la VI seleccionada si mode='1VI'.
        :param fixed_vi_name: Nombre de la VI a fijar si mode='2VIs'.
        :param fixed_descriptor_value: Valor del descriptor a fijar si mode='2VIs' (valor original, sin alias).
        :return: Diccionario {original_key: display_name} de los grupos filtrados.
                 El display_name se ajusta según el modo.
        """
        logger.info(f"Filtrando grupos: mode={mode}, freq={frequency}, primary={primary_vi_name}, fixed_vi={fixed_vi_name}, fixed_desc={fixed_descriptor_value}")
        # Obtener todos los grupos como lista de tuplas (display, key)
        all_groups_tuples = self.get_discrete_analysis_groups(study_id, frequency)
        # Convertir a dict {key: display} para facilitar búsqueda
        all_groups = {key: display for display, key in all_groups_tuples}
        aliases = self.study_service.get_study_aliases(study_id)
        filtered_groups = {}

        if mode == '1VI' and primary_vi_name:
            for key, _ in all_groups.items():
                parts = key.split(';')
                # Solo incluir si tiene exactamente 1 parte y coincide la VI
                if len(parts) == 1:
                    vi_name, descriptor = parts[0].split('=')
                    if vi_name == primary_vi_name:
                        # Display name: solo el descriptor (con alias)
                        alias = aliases.get(descriptor, descriptor)
                        # Mantener VI en display por claridad, pero usar alias
                        filtered_groups[key] = f"{primary_vi_name}: {alias}"
        elif mode == '2VIs' and fixed_vi_name and fixed_descriptor_value:
            fixed_pair_str = f"{fixed_vi_name}={fixed_descriptor_value}"
            for key, _ in all_groups.items():
                parts = key.split(';')
                # Incluir si tiene exactamente 2 partes y una de ellas es el par fijo
                if len(parts) == 2:
                    part1, part2 = parts[0], parts[1]
                    if part1 == fixed_pair_str or part2 == fixed_pair_str:
                        # Encontrar la otra parte (la que varía)
                        other_part = part2 if part1 == fixed_pair_str else part1
                        other_vi_name, other_desc = other_part.split('=')
                        other_alias = aliases.get(other_desc, other_desc)
                        # Nuevo display name: solo la parte que varía (con alias)
                        filtered_groups[key] = f"{other_vi_name}: {other_alias}"

        logger.debug(f"Grupos filtrados resultantes: {filtered_groups}")
        # Devolver ordenado por display name para consistencia en UI
        return dict(sorted(filtered_groups.items(), key=lambda item: item[1]))


    def get_common_columns_for_groups(self, study_id: int, frequency: str, calculation: str, group_keys: list[str]) -> list[str]:
        """
        Encuentra las columnas de datos comunes presentes en las tablas de resumen
        discreto para una combinación específica de frecuencia, cálculo y grupos.

        :param study_id: ID del estudio.
        :param frequency: Frecuencia (ej: 'Cinematica').
        :param calculation: Cálculo (ej: 'Maximo').
        :param group_keys: Lista de claves de grupo (ej: ['CMJ_PRE', 'CMJ_POST']).
        :return: Lista de nombres de columnas comunes ('Attr/Col/Unit').
                 Retorna lista vacía si no hay comunes o falta archivo.
        """
        logger.debug(f"Buscando columnas comunes para estudio {study_id}, "
                     f"freq={frequency}, calc={calculation}, grupos={group_keys}")
        common_columns = None
        tables_path = self.get_discrete_analysis_tables_path(study_id)

        if not tables_path or not group_keys:
            return []

        freq_path = tables_path / frequency
        if not freq_path.exists():
            logger.warning(f"Directorio de frecuencia no encontrado: {freq_path}")
            return []

        for group_key in group_keys:
            # Usar la misma lógica de nombre de archivo seguro que en generate_tables
            safe_group_key_part = group_key.replace('=', '_').replace(';', '__')
            table_filename = f"{calculation}_{frequency}_{safe_group_key_part}.csv"
            table_path = freq_path / table_filename

            if not table_path.exists():
                logger.warning(f"Tabla de resumen no encontrada: {table_path}")
                return [] # Si falta una tabla, no hay columnas comunes

            try:
                # Leer solo las cabeceras para obtener columnas
                # Usar punto decimal '.' para el CSV interno
                df_header = pd.read_csv(table_path, sep=',', decimal='.',
                                        encoding='utf-8', header=[0, 1, 2],
                                        index_col=0, nrows=0)
                logger.debug(f"Cabeceras leídas de {table_filename}: "
                             f"{df_header.columns}")

                # Crear nombres combinados 'Atributo/Columna/Unidad'
                # Añadir manejo de error si attr/col/unit no son strings
                current_columns = set()
                for header_tuple in df_header.columns:
                    try:
                        # Asegurar 3 elementos y convertir a string
                        if len(header_tuple) == 3:
                            # Ya no necesitamos validar ':', asumimos que la generación es correcta
                            attr, col, unit = map(str, header_tuple)
                            current_columns.add(f"{attr}/{col}/{unit}")
                        else:
                            logger.warning(f"Tupla de cabecera con longitud inesperada en "
                                           f"{table_filename}: {header_tuple}. "
                                           f"Se omitirá.")
                    except Exception as e_header:
                        logger.warning(f"Error procesando tupla de cabecera "
                                       f"{header_tuple} en {table_filename}: "
                                       f"{e_header}. Se omitirá.")

                logger.debug(f"Columnas parseadas de {table_filename}: "
                             f"{current_columns}")

                if common_columns is None:
                    common_columns = current_columns
                else:
                    common_columns.intersection_update(current_columns)

                if not common_columns:
                    logger.warning(f"No se encontraron columnas comunes "
                                   f"después de procesar {table_filename}")
                    return []  # Si la intersección es vacía, terminar

            except Exception as e:
                logger.error(f"Error leyendo cabeceras de {table_path}: {e}",
                             exc_info=True)
                return []  # Error al leer una tabla

        if common_columns is None:
            return []

        # Devolver lista ordenada
        return sorted(list(common_columns))

    def perform_individual_analysis(self, study_id: int, config: dict):
        """
        Realiza un análisis individual basado en la configuración, genera un gráfico
        y guarda la configuración.

        :param study_id: ID del estudio.
        :param config: Diccionario con la configuración del análisis:
                       {'name': str, 'frequency': str, 'calculation': str,
                        'column': str, 'groups': list[str],
                        'parametric': bool, 'paired': bool}
        :return: Diccionario con rutas al gráfico y archivo de config generados.
                 {'plot_path': str, 'config_path': str}
        :raises ValueError: Si la configuración es inválida o faltan datos/archivos.
        :raises Exception: Si ocurre un error durante el análisis o graficación.
        """
        analysis_name_log = config.get('name', 'N/A')
        logger.info(f"Iniciando análisis individual para estudio {study_id}: "
                    f"{analysis_name_log}")

        # --- Validación de Configuración ---
        required_keys = ['name', 'frequency', 'calculation', 'column',
                         'groups', 'parametric', 'paired']
        if not all(key in config for key in required_keys):
            raise ValueError("Configuración de análisis incompleta.")
        if len(config['groups']) < 2:
            raise ValueError("Se requieren al menos dos grupos para comparar.")
        if not config['name'] or not config['name'].strip():
            raise ValueError("El nombre del análisis no puede estar vacío.")
        # Validar caracteres inválidos en el nombre
        analysis_name = config['name'].strip()
        invalid_chars = r'<>:"/\|?*'
        if any(char in analysis_name for char in invalid_chars):
            raise ValueError(f"El nombre del análisis contiene caracteres "
                             f"inválidos: {invalid_chars}")

        # --- Preparar Rutas ---
        study_path = self.file_service._get_study_path(study_id)
        if not study_path:
            raise ValueError(f"No se pudo encontrar ruta del estudio {study_id}.")

        analysis_output_dir = (study_path / "Analisis Discreto" / "Individual" /
                               analysis_name)
        try:
            analysis_output_dir.mkdir(parents=True, exist_ok=True)
        except OSError as e:
            raise OSError(f"No se pudo crear directorio de salida para análisis:"
                          f" {analysis_output_dir}\n{e}")

        plot_path = analysis_output_dir / "boxplot.png"
        config_path = analysis_output_dir / "config.json"

        # --- Leer Datos ---
        frequency = config['frequency']
        calculation = config['calculation']
        column_str = config['column']

        # --- Validación y Parseo de Columna (maneja '/' en unidad) ---
        try:
            # Dividir en máximo 3 partes: Atributo, Columna, Unidad (puede contener '/')
            target_column_parts = column_str.split('/', 2)
            if len(target_column_parts) != 3:
                # Si no hay 3 partes, el formato es incorrecto (faltan al menos dos '/')
                raise ValueError("Formato incorrecto, faltan separadores '/'")
            # La tupla para buscar en el DataFrame sigue siendo (attr, col, unit)
            target_multi_index_col = tuple(target_column_parts)
            logger.debug(f"Columna parseada para análisis: {target_multi_index_col}")
        except Exception as e:
             # Captura el ValueError de arriba u otros errores inesperados
             raise ValueError(f"Formato de columna inválido recibido: "
                              f"'{column_str}'. Se esperaba "
                              f"'Atributo/Columna/Unidad'. Error: {e}")

        data_by_group = []
        group_names = config['groups']
        tables_path = self.get_discrete_analysis_tables_path(study_id)
        freq_path = tables_path / frequency

        # Identificar mapeo archivo -> grupo (no usado directamente aquí, pero útil)
        # files_to_groups, _ = self._identify_study_groups(study_id, frequency)

        for group_key in group_names:
            # --- Usar la misma lógica de nombre seguro que en generate_tables ---
            safe_group_key_part = group_key.replace('=', '_').replace(';', '__')
            table_filename = f"{calculation}_{frequency}_{safe_group_key_part}.csv"
            table_path = freq_path / table_filename
            if not table_path.exists():
                # Log detallado del archivo buscado y el error
                logger.error(f"No se encontró tabla resumen requerida. Buscando: {table_path}")
                # Intentar listar archivos en el directorio para depuración
                try:
                    existing_files = [f.name for f in freq_path.iterdir() if f.is_file()]
                    logger.debug(f"Archivos existentes en {freq_path}: {existing_files}")
                except Exception as list_e:
                    logger.debug(f"No se pudo listar archivos en {freq_path}: {list_e}")
                raise FileNotFoundError(f"No se encontró tabla resumen requerida: {table_path}")

            try:
                # Usar punto decimal '.' para leer el CSV interno
                df = pd.read_csv(table_path, sep=',', decimal='.', encoding='utf-8', header=[0, 1, 2], index_col=0)
                # Verificar si la columna existe
                if target_multi_index_col not in df.columns:
                    raise ValueError(f"La columna '{config['column']}' no se "
                                     f"encontró en la tabla {table_filename}")

                # Extraer la serie de datos y filtrar NaNs
                group_data = df[target_multi_index_col].dropna().tolist()

                # --- Manejo de Datos Pareados ---
                # TODO: Implementar lógica de emparejamiento robusta si es pareado.
                # Por ahora, se asume orden consistente de índices (ARCHIVO).
                if config['paired']:
                    logger.warning("Manejo de datos pareados asume índices "
                                   "consistentes. Requiere verificación.")
                    pass  # Continuar con group_data tal cual

                if not group_data:
                    logger.warning(f"No se encontraron datos válidos para grupo "
                                   f"'{group_key}' y columna '{config['column']}'"
                                   f" en {table_filename}")
                    # Añadir lista vacía para mantener correspondencia
                    data_by_group.append([])
                else:
                    data_by_group.append(group_data)

            except Exception as e:
                logger.error(f"Error procesando tabla {table_path} para grupo "
                             f"{group_key}: {e}", exc_info=True)
                raise ValueError(f"Error leyendo datos para grupo {group_key}: {e}")

        # Verificar si tenemos datos para graficar
        if not any(data_by_group):
            raise ValueError("No se encontraron datos válidos en ninguna tabla "
                             "para los grupos y columna seleccionados.")

        # --- Generar Título y Leyendas Específicos del Modo ---
        aliases = self.study_service.get_study_aliases(study_id)
        mode = config.get("grouping_mode")
        primary_vi = config.get("primary_vi_name")
        fixed_vi = config.get("fixed_vi_name")
        fixed_desc_display = config.get("fixed_descriptor_display") # Display name con alias
        calculation = config['calculation'] # Obtener cálculo
        column = config['column'] # Obtener columna completa (Attr/Col/Unit)
        groups = config['groups'] # Obtener claves originales seleccionadas
        frequency = config['frequency'] # Obtener frecuencia

        # Título base
        title = f"{calculation.capitalize()} de {column}"
        group_legend_names = [] # Nombres para la leyenda del gráfico
        group_xaxis_labels = [] # Etiquetas cortas para eje X ("G1", "G2", ...)

        # Obtener display names filtrados (como se mostraron en el diálogo)
        # Necesitamos volver a llamar a get_filtered_discrete_analysis_groups
        # o pasar los display names en la config (mejor opción)
        # Por ahora, reconstruimos la lógica aquí:
        filtered_groups_display = {}
        if mode == '1VI' and primary_vi:
             title += f" ({primary_vi})"
             for key in groups: # groups son las claves originales seleccionadas
                 vi_name, descriptor = key.split('=')
                 if vi_name == primary_vi:
                     alias = aliases.get(descriptor, descriptor)
                     filtered_groups_display[key] = f"{alias}" # Leyenda corta
        elif mode == '2VIs' and fixed_vi and fixed_desc_display:
             title += f" ({fixed_desc_display})"
             fixed_desc_original = fixed_desc_display.split(" (")[0]
             fixed_pair_str = f"{fixed_vi}={fixed_desc_original}"
             for key in groups:
                 parts = key.split(';')
                 if len(parts) == 2:
                     part1, part2 = parts[0], parts[1]
                     other_part = part2 if part1 == fixed_pair_str else part1
                     other_vi_name, other_desc = other_part.split('=')
                     other_alias = aliases.get(other_desc, other_desc)
                     filtered_groups_display[key] = f"{other_vi_name}: {other_alias}" # Leyenda con parte variable
        else: # Fallback
             title += " (Comparación)"
             # Obtener todos los display names (tuplas)
             all_display_groups_tuples = self.get_discrete_analysis_groups(study_id, frequency)
             # Convertir a dict {key: display}
             all_display_groups = {key: display for display, key in all_display_groups_tuples}
             for key in groups:
                 # Extraer solo el nombre base del display name completo ("Grupo X - Nombre Base")
                 full_display = all_display_groups.get(key, key)
                 base_display = full_display.split(" - ", 1)[1] if " - " in full_display else full_display
                 filtered_groups_display[key] = base_display # Usar nombre base como fallback

        # Generar leyendas y etiquetas X ordenadas según las claves originales
        sorted_groups = sorted(groups) # Ordenar claves originales para consistencia
        for i, key in enumerate(sorted_groups):
             group_legend_names.append(filtered_groups_display.get(key, f"Grupo {i+1}"))
             group_xaxis_labels.append(f"G{i+1}") # Usar G1, G2... para eje X

        # Reordenar data_by_group para que coincida con sorted_groups
        # Crear un diccionario temporal para mapear clave original a datos
        # group_names aquí son las claves originales leídas de las tablas, deben coincidir con 'groups'
        data_map = {key: data for key, data in zip(group_names, data_by_group)}
        ordered_data_by_group = [data_map[key] for key in sorted_groups if key in data_map]

        # Obtener ylabel del gráfico (unidad)
        try:
            chart_ylabel = f"{calculation.capitalize()} ({target_column_parts[2]})"
        except IndexError:
            chart_ylabel = f"{calculation.capitalize()}" # Fallback si no hay unidad


        logger.info(f"Título del gráfico: {title}")
        logger.info(f"Etiquetas Eje X: {group_xaxis_labels}")
        logger.info(f"Leyendas del gráfico: {group_legend_names}")


        # --- Realizar Análisis Estadístico ---
        stats_results = None
        if stats:  # Verificar si scipy.stats está disponible
                try:
                    n_groups = len(data_by_group)
                    is_paired = config['paired']
                    is_parametric = config['parametric']
                    test_name = "N/A"
                    p_value = np.nan

                    if n_groups == 2:
                        group1_data = np.array(data_by_group[0])
                        group2_data = np.array(data_by_group[1])

                        if is_paired:
                            # Asegurar misma longitud para tests pareados
                            min_len = min(len(group1_data), len(group2_data))
                            if min_len < 1:
                                raise ValueError("Datos insuficientes para "
                                                 "test pareado.")
                            group1_data = group1_data[:min_len]
                            group2_data = group2_data[:min_len]

                            if is_parametric:
                                test_name = "T-test relacionado"
                                _, p_value = stats.ttest_rel(
                                    group1_data, group2_data, nan_policy='omit'
                                )
                            else:
                                test_name = "Wilcoxon signed-rank"
                                # Wilcoxon requiere > 0 diferencias
                                try:
                                    _, p_value = stats.wilcoxon(
                                        group1_data, group2_data,
                                        nan_policy='omit'
                                    )
                                except ValueError as e:
                                    logger.warning(f"No se pudo ejecutar Wilcoxon "
                                                   f"para {analysis_name_log}: {e}")
                                    p_value = np.nan  # Marcar como no calculable
                        else:  # Independiente
                            if is_parametric:
                                test_name = "T-test independiente"
                                # Log data before test
                                logger.debug(f"Datos para {test_name} (Grupo 1, n={len(group1_data)}): {group1_data[:10]}...") # Muestra primeros 10
                                logger.debug(f"Datos para {test_name} (Grupo 2, n={len(group2_data)}): {group2_data[:10]}...") # Muestra primeros 10
                                # Welch's t-test por defecto
                                try:
                                    stat_result, p_value = stats.ttest_ind(
                                        group1_data, group2_data, equal_var=False,
                                        nan_policy='omit'
                                    )
                                    logger.debug(f"{test_name} resultado: stat={stat_result}, p={p_value}")
                                except Exception as ttest_e:
                                    logger.warning(f"Error ejecutando {test_name} para {analysis_name_log}: {ttest_e}")
                                    p_value = np.nan # Marcar como no calculable si falla
                            else:
                                test_name = "Mann-Whitney U"
                                # Log data before test
                                logger.debug(f"Datos para {test_name} (Grupo 1, n={len(group1_data)}): {group1_data[:10]}...")
                                logger.debug(f"Datos para {test_name} (Grupo 2, n={len(group2_data)}): {group2_data[:10]}...")
                                _, p_value = stats.mannwhitneyu(
                                    group1_data, group2_data,
                                    alternative='two-sided', nan_policy='omit'
                                )

                    elif n_groups > 2:
                        # Filtrar grupos vacíos antes de pasar a ANOVA/Kruskal
                        valid_data_for_test = [np.array(g) for g in data_by_group
                                               if len(g) > 0]
                        if len(valid_data_for_test) < 2:
                            raise ValueError("Se necesitan al menos dos grupos "
                                             "con datos para comparar.")

                        if is_paired:  # ANOVA medidas repetidas / Friedman
                            # Nota: ANOVA medidas repetidas es más complejo.
                            # Usar Friedman como alternativa no paramétrica.
                            if is_parametric:
                                test_name = "ANOVA MR (NO IMPL.)"
                                logger.warning(f"{test_name} para {analysis_name_log}. "
                                               f"Usando Friedman como fallback.")
                                try:
                                    _, p_value = stats.friedmanchisquare(
                                        *valid_data_for_test
                                    )
                                    test_name = "Friedman (fallback)"
                                except ValueError as e:
                                    logger.warning(f"No se pudo ejecutar Friedman "
                                                   f"para {analysis_name_log}: {e}")
                                    p_value = np.nan
                            else:
                                test_name = "Friedman"
                                try:
                                    _, p_value = stats.friedmanchisquare(
                                        *valid_data_for_test
                                    )
                                except ValueError as e:
                                    logger.warning(f"No se pudo ejecutar Friedman "
                                                   f"para {analysis_name_log}: {e}")
                                    p_value = np.nan
                        else:  # ANOVA de un factor / Kruskal-Wallis
                            if is_parametric:
                                test_name = "ANOVA (un factor)"
                                _, p_value = stats.f_oneway(*valid_data_for_test)
                            else:
                                test_name = "Kruskal-Wallis"
                                _, p_value = stats.kruskal(*valid_data_for_test,
                                                           nan_policy='omit')

                    if not np.isnan(p_value):
                        stats_results = {'test_name': test_name,
                                         'p_value': p_value}
                        logger.info(f"Análisis estadístico para {analysis_name_log}: "
                                    f"{test_name}, p-valor = {p_value:.4f}")
                    else:
                        logger.warning(f"No se pudo calcular p-valor para "
                                       f"{analysis_name_log} con test {test_name}.")

                except ValueError as ve:
                    logger.error(f"Error en datos para análisis estadístico de "
                                 f"{analysis_name_log}: {ve}")
                except Exception as e_stat:
                    logger.error(f"Error inesperado durante análisis estadístico "
                                 f"de {analysis_name_log}: {e_stat}",
                                 exc_info=True)
        else:
            logger.warning("Scipy no encontrado. Omitiendo pruebas estadísticas.")

        # --- Generar Gráfico ---
        try:
            # Usar título y leyendas generados arriba
            # Pasar los datos ordenados
            charting.create_comparison_boxplot(
                data_by_group=ordered_data_by_group, # Usar datos ordenados
                group_xaxis_labels=group_xaxis_labels, # Etiquetas cortas G1, G2...
                group_legend_names=group_legend_names, # Leyendas específicas del modo
                title=title, # Título específico del modo
                ylabel=chart_ylabel,
                output_path=plot_path,
                stats_results=stats_results
            )
            logger.info(f"Gráfico boxplot generado en: {plot_path}")
        except Exception as e:
            logger.error(f"Error generando gráfico para análisis "
                         f"{analysis_name}: {e}", exc_info=True)
            raise Exception(f"Error generando el gráfico: {e}")

        # --- Guardar Gráfico Interactivo (HTML) ---
        interactive_plot_path = analysis_output_dir / "boxplot_interactive.html"
        try:
            # Pasar datos ordenados y etiquetas/leyendas correctas
            charting.create_interactive_comparison_boxplot(
                data_by_group=ordered_data_by_group, # Usar datos ordenados
                group_xaxis_labels=group_xaxis_labels, # Etiquetas cortas G1, G2...
                group_legend_names=group_legend_names, # Leyendas específicas del modo
                title=title, # Título específico del modo
                ylabel=chart_ylabel,
                output_path=interactive_plot_path
            )
        except Exception as e_plotly:
            logger.error(f"Error generando gráfico interactivo para análisis "
                         f"{analysis_name}: {e_plotly}", exc_info=True)
            # No relanzar, el estático ya se generó, pero marcar como no disponible
            interactive_plot_path = None


        # --- Guardar Configuración (incluyendo resultados estadísticos) ---
        config_to_save = config.copy()
        if stats_results:
            # Guardar solo test y p-valor para evitar problemas de serialización
            config_to_save['stats_results'] = {
                'test_name': stats_results.get('test_name', 'N/A'),
                'p_value': stats_results.get('p_value') # Puede ser NaN
            }
        else:
             config_to_save['stats_results'] = None # Indicar que no hubo test

        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                # Usar json.dump con manejo de NaN
                json.dump(config_to_save, f, indent=4, allow_nan=True)
            # Pequeña pausa para asegurar visibilidad del archivo
            time.sleep(0.1)
            logger.info(f"Configuración análisis guardada en: {config_path}")
        except Exception as e:
            logger.error(f"Error guardando config análisis {analysis_name}: {e}", exc_info=True)
            # No relanzar necesariamente, el gráfico ya se generó
            # Considerar eliminar gráfico si falla guardado de config?

        result_paths = {
            'plot_path': str(plot_path),
            'config_path': str(config_path)
        }
        if interactive_plot_path and interactive_plot_path.exists():
            result_paths['interactive_plot_path'] = str(interactive_plot_path)

        return result_paths

    def _get_individual_analysis_base_dir(self, study_id: int) -> Path | None:
        """Obtiene el directorio base para los análisis individuales."""
        study_path = self.file_service._get_study_path(study_id)
        if not study_path:
            logger.error(f"No se pudo encontrar ruta estudio {study_id} para "
                         f"análisis individual.")
            return None
        return study_path / "Analisis Discreto" / "Individual"

    def get_individual_analysis_path(self, study_id: int, analysis_name: str) \
            -> Path | None:
        """Obtiene la ruta completa al directorio de un análisis individual específico."""
        base_dir = self._get_individual_analysis_base_dir(study_id)
        if not base_dir:
            return None
        # Validar nombre por si acaso
        invalid_chars = r'<>:"/\|?*'
        if any(char in analysis_name for char in invalid_chars):
            logger.error(f"Nombre de análisis inválido solicitado: "
                         f"{analysis_name}")
            return None
        return base_dir / analysis_name

    def list_individual_analyses(self, study_id: int) -> list[dict]:
        """
        Lista los análisis individuales guardados para un estudio.

        :param study_id: ID del estudio.
        :return: Lista de diccionarios:
                 {'name': str, 'path': Path, 'config': dict, 'mtime': float,
                  'plot_path': Path, 'interactive_plot_path': Path | None}
        """
        analyses = []
        base_dir = self._get_individual_analysis_base_dir(study_id)
        if not base_dir or not base_dir.exists():
            return []

        for item_path in base_dir.iterdir():
            if item_path.is_dir():
               analysis_name = item_path.name
               config_path = item_path / "config.json"
               plot_path = item_path / "boxplot.png" # Asumir nombre fijo PNG
               interactive_plot_path = item_path / "boxplot_interactive.html" # Asumir nombre fijo HTML

               if config_path.exists() and config_path.is_file():
                    try:
                        with open(config_path, 'r', encoding='utf-8') as f:
                            config_data = json.load(f)
                        # Usar mtime del config.json como referencia
                        mtime = config_path.stat().st_mtime
                        analyses.append({
                            'name': analysis_name,
                            'path': item_path,
                           'config': config_data,
                           'mtime': mtime,
                           'plot_path': plot_path,
                           # Añadir ruta interactiva si existe
                           'interactive_plot_path': interactive_plot_path if interactive_plot_path.exists() else None
                       })
                    except json.JSONDecodeError:
                        logger.error(f"Error leyendo config.json para análisis "
                                     f"'{analysis_name}' en estudio {study_id}.")
                    except Exception as e:
                        logger.error(f"Error procesando análisis "
                                     f"'{analysis_name}' en estudio {study_id}: {e}",
                                     exc_info=True)
               # Mover la advertencia aquí, fuera del try-except de lectura JSON
               # Solo advertir si el archivo config no existe o no es un archivo
               elif not config_path.exists() or not config_path.is_file():
                   logger.warning(f"Directorio análisis '{analysis_name}' "
                                  f"encontrado sin archivo config.json válido en "
                                  f"estudio {study_id}.")

        # Ordenar por fecha de modificación (más reciente primero)
        analyses.sort(key=lambda x: x['mtime'], reverse=True)
        return analyses

    def delete_individual_analysis(self, study_id: int, analysis_name: str):
        """
        Elimina la carpeta y contenido de un análisis individual específico.

        :param study_id: ID del estudio.
        :param analysis_name: Nombre del análisis a eliminar.
        :raises ValueError: Si el nombre del análisis es inválido.
        :raises FileNotFoundError: Si el directorio del análisis no existe.
        :raises OSError: Si ocurre un error al eliminar el directorio.
        """
        analysis_dir = self.get_individual_analysis_path(study_id,
                                                         analysis_name)
        if not analysis_dir:
            # get_individual_analysis_path ya loggea el error
            raise ValueError(f"Nombre de análisis inválido o ruta de estudio "
                             f"no encontrada: {analysis_name}")

        if not analysis_dir.exists():
            raise FileNotFoundError(f"El directorio del análisis no existe: "
                                    f"{analysis_dir}")
        if not analysis_dir.is_dir():
            raise ValueError(f"La ruta del análisis no es un directorio: "
                             f"{analysis_dir}")

        try:
            shutil.rmtree(analysis_dir)
            logger.info(f"Análisis individual eliminado: {analysis_dir}")
            # Opcional: Limpiar directorios padre si quedan vacíos
        except OSError as e:
            logger.error(f"Error eliminando directorio análisis "
                         f"{analysis_dir}: {e}", exc_info=True)
            raise
